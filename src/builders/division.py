"""Sheets 6 & 7: Division-wise TB Incidence (annual) and Cumulative Summary."""

from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook

from src.fetchers.local import DIVISION_CUMULATIVE, DIVISION_INCIDENCE
from src.styles import BORDER, FILLS, autofit, write_data, write_headers, write_note

_DIVISIONS = ["Barishal", "Chattogram", "Dhaka", "Khulna", "Mymensingh", "Rajshahi", "Rangpur", "Sylhet"]

# ── Sheet 6: Annual incidence ─────────────────────────────────────────────────

def _annual_headers() -> list[str]:
    heads = ["Year"]
    for d in _DIVISIONS:
        heads += [f"{d}\nRate*", f"{d}\nCases"]
    heads.append("National\nTotal")
    return heads


def build_annual(wb: Workbook) -> None:
    ws = wb.create_sheet("6. Division-wise Incidence")
    ws.sheet_properties.tabColor = "1F4E79"
    headers = _annual_headers()
    write_headers(ws, headers, "blue")

    for i, row_data in enumerate(DIVISION_INCIDENCE):
        yr = str(row_data[0])
        vals = list(row_data[1:])  # 8 divisions × (rate, cases) + national total
        data_row = i + 2

        write_data(ws, data_row, 1, yr, i, bold=True, align="center")
        col = 2
        for j in range(8):
            rate  = vals[j * 2]
            cases = vals[j * 2 + 1]
            write_data(ws, data_row, col,     rate,  i, num_format="0")
            write_data(ws, data_row, col + 1, cases, i, num_format="#,##0")
            col += 2
        write_data(ws, data_row, col, vals[16], i, num_format="#,##0")

    ws.freeze_panes = "B2"
    note = (
        "* Rate = estimated incidence per 100,000 population. Cases = absolute estimated count. "
        "Mymensingh division separated from Dhaka in 2015; shown as None pre-2015. "
        "Source: DGHS/NTP Bangladesh Annual Reports."
    )
    write_note(ws, len(DIVISION_INCIDENCE) + 3, 1, len(headers), note)
    autofit(ws)


# ── Sheet 7: Cumulative summary ───────────────────────────────────────────────

def build_cumulative(wb: Workbook) -> None:
    ws = wb.create_sheet("7. Division Cumulative")
    ws.sheet_properties.tabColor = "375623"
    write_headers(ws, ["Division", "Cumulative Cases\n(2000–2024)", "Estimated Deaths\n(2000–2024)", "Case Fatality\nRatio (%)"], "green")

    for i, (div, cases, deaths) in enumerate(DIVISION_CUMULATIVE):
        cfr = round(deaths / cases * 100, 1)
        write_data(ws, i + 2, 1, div,   i, bold=True, align="left")
        write_data(ws, i + 2, 2, cases, i, num_format="#,##0")
        write_data(ws, i + 2, 3, deaths,i, num_format="#,##0")
        write_data(ws, i + 2, 4, cfr,   i, num_format="0.0")

    # Total row
    total_row = len(DIVISION_CUMULATIVE) + 2
    tc = sum(r[1] for r in DIVISION_CUMULATIVE)
    td = sum(r[2] for r in DIVISION_CUMULATIVE)
    for col, (val, fmt, align) in enumerate(
        [("Bangladesh Total", None, "left"), (tc, "#,##0", "right"), (td, "#,##0", "right"), (round(td / tc * 100, 1), "0.0", "right")],
        start=1,
    ):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=10)
        cell.fill = FILLS["total"]
        cell.border = BORDER
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            cell.number_format = fmt

    note = (
        "¹ Mymensingh became a separate division in 2015; pre-2015 cases allocated from Dhaka records. "
        "Estimated deaths derived by applying national case-fatality ratio to division case counts."
    )
    write_note(ws, total_row + 2, 1, 4, note)
    autofit(ws)
