"""Shared Excel styling utilities for the TB Bangladesh workbook."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Colour palette ────────────────────────────────────────────────────────────
TAB_COLOURS = {
    "blue":   "1F4E79",
    "teal":   "2E75B6",
    "green":  "375623",
    "purple": "7030A0",
    "red":    "833C00",
    "orange": "C55A11",
    "grey":   "595959",
}

FILLS = {
    **{name: PatternFill("solid", fgColor=colour) for name, colour in TAB_COLOURS.items()},
    "row_even": PatternFill("solid", fgColor="D9E1F2"),
    "row_odd":  PatternFill("solid", fgColor="FFFFFF"),
    "total":    PatternFill("solid", fgColor="C6EFCE"),
    "warn":     PatternFill("solid", fgColor="FFF2CC"),
}

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── Cell writers ──────────────────────────────────────────────────────────────
def write_header(
    ws: Worksheet,
    row: int,
    col: int,
    value: str,
    colour: str = "blue",
    row_height: int | None = 45,
) -> None:
    if row_height is not None:
        ws.row_dimensions[row].height = row_height
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = FILLS[colour]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def write_data(
    ws: Worksheet,
    row: int,
    col: int,
    value: object,
    stripe: int = 0,
    bold: bool = False,
    num_format: str | None = None,
    align: str = "right",
    warn: bool = False,
    total: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    if total:
        cell.fill = FILLS["total"]
    elif warn:
        cell.fill = FILLS["warn"]
    else:
        cell.fill = FILLS["row_even"] if stripe % 2 == 0 else FILLS["row_odd"]
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if num_format:
        cell.number_format = num_format


def write_note(ws: Worksheet, row: int, start_col: int, end_col: int, text: str) -> None:
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = Font(italic=True, size=9, color="595959")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    ws.row_dimensions[row].height = 40


def write_headers(ws: Worksheet, headers: list[str], colour: str = "blue") -> None:
    for col, header in enumerate(headers, 1):
        write_header(ws, 1, col, header, colour)


# ── Column auto-width ─────────────────────────────────────────────────────────
def autofit(ws: Worksheet, min_width: int = 10, max_width: int = 28) -> None:
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=0) + 3
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(width, min_width), max_width
        )


# ── Safe value extraction ─────────────────────────────────────────────────────
def safe(row: dict, key: str, typ: type = float) -> float | int | None:
    """Extract and cast a value from a WHO CSV row dict; return None if missing."""
    val = row.get(key, "")
    if val in ("", None):
        return None
    try:
        return typ(val)
    except (ValueError, TypeError):
        return None
